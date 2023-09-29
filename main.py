import re
import os
import configparser
from typing import Optional

from openpyxl.styles import PatternFill
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils import Logger

config = configparser.ConfigParser()

with open("./settings/settings.ini", "r") as file:
    config.read_file(file)

INPUT_PATH = config.get("paths", "input")

MASTER_PATH = config.get("paths", "master")

class UnreviewedToMaster:
    """Moves matching unreviewed products to master"""
    def __init__(self) -> None:
        self.logger = Logger(__class__.__name__)
        self.logger.info("*****UnreviewedToMaster started*****")

        self.qualified_row, self.notes_row = 8, 9

        self.green = "91bf4d"

        self.processed_extensions = ["_added", "_nothing"]

        self.input_files = self.__read_filenames_in_dir()

        self.master_path, self.master_wb, self.master_ws = self.__get_master_workbook()

    def __read_filenames_in_dir(self) -> Optional[list[str]]:
        """Reads the filenames that exists in input directory"""
        try:
            self.logger.info("Reading the input path...")

            files = [f"{INPUT_PATH}{file}" for file in os.listdir(INPUT_PATH) if file.endswith(".xlsx")]

            unprocessed_files = [file for file in files if not any(
                                 re.search(rf"{extension}", file) for extension in self.processed_extensions)]

            self.logger.info("%s input files found." % len(unprocessed_files))

            return unprocessed_files
        
        except:
            self.logger.error("Fatal error while reading the input directory!")
        
        self.logger.warn("No files found in the input path!")
    
    def __get_master_workbook(self) -> Optional[tuple[str, Workbook, Worksheet]]:
        """Reads the master excel path to get the master file"""
        self.logger.info("Reading the master file path...")

        for file in os.listdir(MASTER_PATH):
            if file.endswith(".xlsx") and re.search(r"master", file, re.I):
                wb = self.__load_workbook(f"{MASTER_PATH}{file}")

                ws = self.__locate_target_worksheet(wb, "unreviewed")

                if ws is not None:
                    return f"{MASTER_PATH}{file}", wb, ws
                
                self.logger.error("'Unreviewed matching products' sheet not found in master!")
        
        self.logger.error("No master file found in master path!")

    def __load_workbook(self, file_path: str) -> Workbook:
        """Loads an excel file for processing"""
        wb = load_workbook(file_path, rich_text=True)

        return wb

    def __locate_target_worksheet(self, wb: Workbook, ws_name: Optional[str]=None) -> Optional[Worksheet]:
        """Locates the target worksheet in the excel file"""
        if ws_name is None:
            return wb[wb.sheetnames[0]]

        else:
            for sheet in wb.sheetnames:
                if re.search(rf"{ws_name}", sheet, re.I):
                    return wb[sheet]
    
    def __process_rows(self, ws: Worksheet) -> str:
        """Processes rows in the given input worksheet"""
        rows_added, header_skipped, extension = 0, False, "_nothing"

        start_row = self.master_ws.max_row + 1

        for row in ws.rows:
            if not header_skipped:
                header_skipped = True 
                continue

            values = [cell.value for cell in row]

            qualified, notes = values[self.qualified_row], values[self.notes_row]

            if qualified == 'Qualified?' or notes == 'Notes': continue

            if qualified is not None or notes is not None:
                rows_added += 1
                
                self.master_ws.append(values)
        
        if rows_added > 0:
            extension = "_added"

            self.__style_links_and_roi(start_row)
                        
            self.logger.info("Rows added to master: {}".format(rows_added))

            self.master_wb.save(self.master_path)
        
        return extension
    
    def __style_links_and_roi(self, start_row: int) -> None:
        """Styles the hyperlinks and ROI column"""
        for row in self.master_ws.iter_rows(max_col=self.master_ws.max_column, min_row=start_row):
            for cell in row:
                if str(cell.internal_value).startswith("=HY"):
                    cell.style = "Hyperlink"

                    continue
                
                if str(cell.internal_value).endswith("%"):
                    cell.fill = PatternFill(fill_type="solid", end_color=self.green, start_color=self.green)
    
    def __save_input_file(self, wb: Workbook, filepath: str, original_file: str) -> None:
        """Saves the input file"""
        if os.path.isfile(original_file):
            os.remove(original_file)

        wb.save(filepath)

    def run(self) -> None:
        """Entry point to the unreviewed rows processor"""
        self.logger.info("Processing input files...")

        for file in self.input_files:
            self.logger.info("Processing file >> {}".format(file))

            wb = self.__load_workbook(file)

            ws = self.__locate_target_worksheet(wb)

            extension = self.__process_rows(ws)

            file_path = file.replace(".xlsx", f"{extension}.xlsx")

            self.__save_input_file(wb, file_path, file)
        
        self.logger.info("Done.")

if __name__ == "__main__":
    app = UnreviewedToMaster()
    app.run()